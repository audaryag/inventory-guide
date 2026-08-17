#!/usr/bin/env python3
"""Writes 'Variables and Calculations.xlsx' in the shapes the report reads.

Every sheet carries the exact headers the queries look for, a frozen header row and
a read-me tab. The three plants are filled in, the MW sheet is the new month-per-column
layout, and the masters are headers only: their rows are the user's own data and are
pasted in, never invented here.
"""
import datetime
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10)
FILL = PatternFill("solid", fgColor="14532D")
NOTE = Font(name="Arial", size=10, italic=True, color="4B5A50")

MONTHS = [datetime.date(2026, 3, 31), datetime.date(2026, 4, 30),
          datetime.date(2026, 5, 31), datetime.date(2026, 6, 30)]

PLANTS = [("1902", "Jaipur Module", 1),
          ("1900", "Dholera Module", 2),
          ("1905", "Dholera Cell", 3)]

TECHS = ["G12 Perc Module", "G12R Topcon Module", "M10 Perc Module",
         "M10 Topcon Module", "M10R Perc Module", "M10R Topcon Module"]


def sheet(wb, title, headers, widths, rows=(), note=None, dates_from=None):
    ws = wb.create_sheet(title)
    ws.append(list(headers))
    for c in ws[1]:
        c.font, c.fill = HEAD, FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in rows:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = BODY
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if dates_from:
        for c in ws[1][dates_from:]:
            c.number_format = "dd-mmm-yy"
    ws.freeze_panes = "A2"
    if note:
        # two columns clear of the last heading: a note typed under the headings would be
        # read as a row of data, and a sentence in the plant column is a plant
        c = ws.cell(row=1, column=len(list(headers)) + 2, value=note)
        c.font = NOTE
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(len(list(headers)) + 2)].width = 60
    return ws


def readme(wb):
    lines = [
        ("Variables and Calculations - what each sheet is, and the one rule for each", True),
        ("", False),
        ("This workbook is the report's master. The stock and trial balance folders are read", False),
        ("as they come out of SAP and are never written to; everything the report has to be", False),
        ("told rather than measured lives here.", False),
        ("", False),
        ("Sheet names and column headings are matched loosely - case, spaces and punctuation", False),
        ("are ignored, and several spellings are accepted - but keep them as they are here and", False),
        ("nothing has to be guessed.", False),
        ("", False),
        ("MW Capacity   one row per plant (and per technology if you have it), one column per", False),
        ("              month. The column heading is a real date: the month-end its figures", False),
        ("              take effect from. A NEW MONTH IS A NEW COLUMN - never overwrite a", False),
        ("              figure already typed, because a month with no column of its own keeps", False),
        ("              the last figure to its left, so overwriting rewrites past months.", False),
        ("              An empty cell means unchanged; a dash means nought. Leave the Techno", False),
        ("              column out and the row is that plant's whole capacity, which gives", False),
        ("              days of cover per plant and leaves it blank per technology.", False),
        ("", False),
        ("Plant Master  the three plants and the order they read in. The names themselves are", False),
        ("              fixed inside the report - 1902 Jaipur Module, 1900 Dholera Module,", False),
        ("              1905 Dholera Cell - so a row typed here the wrong way round cannot", False),
        ("              rename a plant on any page. A fourth plant added here does appear.", False),
        ("", False),
        ("Constants     one row per number the report is told rather than measures. Add a new", False),
        ("              row with the date it takes effect; do not overwrite. RM_MW_FACTOR is", False),
        ("              needed for RM megawatts and is 580 unless you change it.", False),
        ("", False),
        ("RM Nature     one row per raw-material code: its nature, the group it belongs to", False),
        ("              (Module or Cell, which is what draws the two blocks on the RM page)", False),
        ("              and its BOM standard quantity. Paste your rows under the headings.", False),
        ("", False),
        ("FG Master     one row per finished-goods code and the technology it is.", False),
        ("", False),
        ("TB Master     one row per GL account: what it is, and the plant it belongs to. The", False),
        ("              GL number is the only key the trial balance and this sheet share, so", False),
        ("              an account with no plant here is an account the report cannot place -", False),
        ("              write 1902, 1900 or 1905 against it, or the plant's name.", False),
        ("", False),
        ("One habit to keep: close this workbook before refreshing the report.", False),
    ]
    ws = wb.create_sheet("Read me", 0)
    for i, (line, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name="Arial", size=10, bold=bold)
    ws.column_dimensions["A"].width = 96
    return ws


def build(path):
    wb = Workbook()
    wb.remove(wb.active)
    readme(wb)

    sheet(wb, "MW Capacity", ["Techno", "Plant"] + MONTHS,
          [26, 10] + [13] * len(MONTHS),
          [(t, code) for t in TECHS for code, _n, _s in PLANTS],
          note="A new month is a new column to the right, headed with its date. "
               "Leave a cell empty to mean unchanged; a dash means nought.",
          dates_from=2)

    sheet(wb, "Plant Master", ["Valuation Area", "Plant Name", "Sort"],
          [16, 24, 8],
          [(code, name, srt) for code, name, srt in PLANTS],
          note="The report fixes these three names itself; this sheet sets the order "
               "they read in, and is where a fourth plant would be added.")

    sheet(wb, "Constants", ["Effective From", "Constant Name", "Value"],
          [16, 26, 12],
          [(datetime.date(2025, 4, 1), "RM_MW_FACTOR", 580)],
          note="Add a new row with its effective date rather than editing a number: "
               "editing one rewrites every month it was already used in.")
    ws = wb["Constants"]
    ws["A2"].number_format = "dd-mmm-yy"

    sheet(wb, "RM Nature",
          ["Valuation Area", "Material", "Material Description", "Nature",
           "Group Nature", "BOM Std Qty", "Item"],
          [16, 18, 40, 22, 16, 14, 10],
          note="Paste your rows under these headings. Group Nature is what draws the "
               "Module and Cell blocks on the RM page: write Module or Cell.")

    sheet(wb, "FG Master",
          ["Valuation Area", "Material", "Material Description", "Nature"],
          [16, 18, 40, 26],
          note="Nature is the technology - the same words as the Techno column on "
               "MW Capacity, so days of cover can find its capacity.")

    sheet(wb, "TB Master",
          ["GL Account Number", "GL Account Description", "Nature", "Plant", "Sort Order"],
          [20, 44, 26, 12, 12],
          note="Plant takes 1902, 1900 or 1905 - or the plant's name. An inventory GL "
               "with this column empty is one the trial balance cannot place, which is "
               "what Checks reports as an unresolved profit centre.")

    wb.save(path)
    return path


if __name__ == "__main__":
    out = pathlib.Path(sys.argv[1] if len(sys.argv) > 1
                       else "/home/ubuntu/Variables and Calculations.xlsx")
    print("wrote", build(out))
