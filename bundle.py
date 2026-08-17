#!/usr/bin/env python3
"""Builds the download zip: two project variants, a Tabular Editor script, a pRoot
setter, and the raw query/measure text. No instructions inside the zip — those live
on the Auto tab of the guide.

    python3 pbip.py            # writes /home/ubuntu/pbip        (full, 55 visuals)
    PBIP_EMPTY=1 python3 pbip.py   # writes /home/ubuntu/pbip-empty  (model only)
    python3 bundle.py          # zips both into InventoryReport-pbip.zip
"""
import json, pathlib, shutil, subprocess, sys, zipfile

HERE = pathlib.Path(__file__).parent
sys.path.insert(0, str(HERE))
import build  # noqa: E402

GUIDE = pathlib.Path("/home/ubuntu/BUILD_GUIDE.md")
STAGE = pathlib.Path("/home/ubuntu/bundle")
ZIP = HERE / "InventoryReport-pbip.zip"
# bumped on every published fix, and used as the folder name inside the zip so an old
# extraction can never be mistaken for a new one
BUILD = 30

md = GUIDE.read_text()
queries = build.parse_queries(md[md.index("# Appendix A"):md.index("# Appendix B")])
measures = build.parse_measures(md[md.index("# Appendix B"):])


def csx():
    """Tabular Editor 2 C# script: creates every measure in one paste."""
    lines = ["// Tabular Editor 2 -> C# Script tab -> paste -> F5, then Ctrl+S.",
             "// Creates or overwrites every report measure on factInventory.",
             'var t = Model.Tables["factInventory"];',
             "int made = 0, updated = 0;"]
    for m in measures:
        expr = m["code"].split("=", 1)[1].strip()
        name = m["name"].replace('"', '\\"')
        lit = '@"' + expr.replace('"', '""') + '"'
        lines += [f'if (t.Measures.Contains("{name}")) '
                  f'{{ t.Measures["{name}"].Expression = {lit}; updated++; }}',
                  f'else {{ var m = t.AddMeasure("{name}", {lit}); '
                  f'm.DisplayFolder = "Report measures"; made++; }}']
    lines += ['Info(made + " measures created, " + updated + " updated.");']
    return "\n".join(lines) + "\n"


PS1 = r'''# Sets pRoot inside an extracted project folder, so nothing is typed in Power BI.
#   Right-click this file -> Run with PowerShell, or:
#   powershell -ExecutionPolicy Bypass -File .\set-proot.ps1 -Root "C:\Users\me\Desktop\Inventory Report"
param([Parameter(Mandatory=$true)][string]$Root)

if (-not (Test-Path -LiteralPath $Root)) { throw "That folder does not exist: $Root" }
foreach ($sub in 'RM Raw','FG Raw','Consble Raw','TB') {
  if (-not (Test-Path -LiteralPath (Join-Path $Root $sub))) {
    Write-Warning "No '$sub' sub-folder inside $Root - refresh will fail until it exists."
  }
}
$Root = $Root.TrimEnd('\')
$here = Split-Path -Parent $MyInvocation.MyCommand.Path
$hits = Get-ChildItem -LiteralPath $here -Recurse -Filter model.bim
if (-not $hits) { throw "No model.bim found under $here - extract the zip first." }

foreach ($f in $hits) {
  $json = Get-Content -LiteralPath $f.FullName -Raw | ConvertFrom-Json
  $p = $json.model.expressions | Where-Object { $_.name -eq 'pRoot' }
  if (-not $p) { Write-Warning "no pRoot in $($f.FullName)"; continue }
  $esc = $Root -replace '\\', '\\'
  $p.expression = '"' + $esc + '" meta [IsParameterQuery=true, Type="Text", IsParameterQueryRequired=true]'
  $json | ConvertTo-Json -Depth 100 | Set-Content -LiteralPath $f.FullName -Encoding UTF8
  Write-Host "pRoot set to $Root in $($f.Directory.Name)"
}
Write-Host "Done. Open the .pbip and press Refresh."
'''


def mw_template(path):
    """The MW Capacity sheet in the layout the report now reads: a plant per row, a
    month per column. A new month is a new column, so nothing already typed is touched -
    each date column is the date its figures take effect from."""
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "MW Capacity"
    head = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body = Font(name="Arial", size=10)
    fill = PatternFill("solid", fgColor="14532D")
    months = [datetime.date(2026, 3, 31), datetime.date(2026, 4, 30),
              datetime.date(2026, 5, 31), datetime.date(2026, 6, 30)]

    ws.append(["Techno", "Plant"] + months)
    for c in ws[1]:
        c.font, c.fill = head, fill
        c.alignment = Alignment(horizontal="center")
    for tech in ["G12 Perc Module", "G12R Topcon Module", "M10 Perc Module",
                 "M10 Topcon Module", "M10R Perc Module", "M10R Topcon Module"]:
        for code in ["1902", "1900", "1905"]:
            ws.append([tech, code])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = body
    for col, w in zip("ABCDEF", [26, 10, 13, 13, 13, 13]):
        ws.column_dimensions[col].width = w
    for c in ws[1][2:]:
        c.number_format = "dd-mmm-yy"

    n = wb.create_sheet("Read me")
    for i, line in enumerate([
        "MW Capacity \u2014 how the report reads this sheet",
        "",
        "One row per plant, one column per month. The column heading is a real date:",
        "the month-end the figures in that column take effect from.",
        "",
        "A new month is a NEW COLUMN on the right. Never overwrite a figure already",
        "typed: a month with no column of its own keeps the last figure to its left,",
        "so overwriting rewrites history.",
        "",
        "The Techno column is optional. Leave it out and the row is that plant's whole",
        "capacity; the report then shows days of cover per plant and leaves days per",
        "technology blank, which is honest rather than invented.",
        "",
        "An empty cell means unchanged since the last column. A dash means nought.",
        "Plant codes: 1902 Jaipur Module, 1900 Dholera Module, 1905 Dholera Cell.",
    ], 1):
        c = n.cell(row=i, column=1, value=line)
        c.font = Font(name="Arial", size=10, bold=(i == 1))
    n.column_dimensions["A"].width = 84
    wb.save(path)


def main():
    if STAGE.exists():
        shutil.rmtree(STAGE)
    STAGE.mkdir(parents=True)

    subprocess.run([sys.executable, "pbip.py"], cwd=HERE, check=True,
                   stdout=subprocess.DEVNULL)
    shutil.copytree("/home/ubuntu/pbip", STAGE / "1 - full report")
    subprocess.run([sys.executable, "pbip.py"], cwd=HERE, check=True,
                   env={"PBIP_LEGACY": "1", "PATH": "/usr/bin:/bin"},
                   stdout=subprocess.DEVNULL)
    shutil.copytree("/home/ubuntu/pbip-legacy", STAGE / "2 - older power bi")

    subprocess.run([sys.executable, "pbip.py"], cwd=HERE, check=True,
                   env={"PBIP_EMPTY": "1", "PATH": "/usr/bin:/bin"},
                   stdout=subprocess.DEVNULL)
    shutil.copytree("/home/ubuntu/pbip-empty", STAGE / "3 - model only")

    (STAGE / "4 - tabular editor").mkdir()
    (STAGE / "4 - tabular editor" / "add-all-measures.csx").write_text(csx())

    mw_template(STAGE / "MW Capacity - sheet layout.xlsx")

    (STAGE / "set-proot.ps1").write_text(PS1)
    shutil.copy(HERE / "inventory-theme.json", STAGE / "inventory-theme.json")

    q = STAGE / "5 - plain text" / "queries"
    m = STAGE / "5 - plain text" / "measures"
    q.mkdir(parents=True)
    m.mkdir(parents=True)
    for i, item in enumerate(queries, 1):
        (q / f"{i:02d} {item['name']}.m").write_text(item["code"] + "\n")
    for i, item in enumerate(measures, 1):
        safe = item["name"].replace("/", "-").replace("%", "pct")
        (m / f"{i:02d} {safe}.dax").write_text(item["code"] + "\n")

    if ZIP.exists():
        ZIP.unlink()
    with zipfile.ZipFile(ZIP, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(STAGE.rglob("*")):
            if f.is_file():
                z.write(f, f"InventoryReport build {BUILD}/"
                           + f.relative_to(STAGE).as_posix())
    n = len(zipfile.ZipFile(ZIP).namelist())
    print(f"wrote build {BUILD} {ZIP} ({ZIP.stat().st_size // 1024} KB, {n} files): "
          f"{len(queries)} queries, {len(measures)} measures")


if __name__ == "__main__":
    main()
